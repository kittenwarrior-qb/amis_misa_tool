"""
converter.py – Logic chuyển đổi dữ liệu Bravo → MISA

Không chứa bất kỳ code giao diện nào.
Hàm convert() là điểm duy nhất được gọi từ bên ngoài.
"""

import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook

# ─────────────────────────────────────────────────────────────────────────────
# Cấu hình vị trí dữ liệu trong file Bravo
# ─────────────────────────────────────────────────────────────────────────────

BRAVO_HEADER_ROW = 4   # Dòng chứa tên cột (1-indexed)
BRAVO_DATA_ROW   = 5   # Dòng đầu tiên có dữ liệu (1-indexed)
BATCH_SIZE       = 1000

# ─────────────────────────────────────────────────────────────────────────────
# Headers của file output (lấy từ template MISA)
# ─────────────────────────────────────────────────────────────────────────────

CUSTOMER_HEADERS = [
    "Mã khách hàng (*)",
    "Tên khách hàng (*)",
    "Tên viết tắt",
    "Email",
    "Điện thoại",
    "Số CMND/CCCD",
    "Nguồn gốc",
    "Mã số thuế",
    "Lĩnh vực",
    "Loại khách hàng",
    "Ngành nghề",
    "Loại hình",
    "Nhân viên bán hàng",
    "Quốc gia (Hóa đơn)",
    "Tỉnh/Thành phố (Hóa đơn)",
    "Quận/Huyện (Hóa đơn)",
    "Phường/Xã (Hóa đơn)",
    "Số nhà, Đường phố (Hóa đơn)",
    "Mã vùng (Hóa đơn)",
    "Địa chỉ (Hóa đơn)",
    "Quốc gia (Giao hàng)",
    "Tỉnh/Thành phố (Giao hàng)",
    "Quận/Huyện (Giao hàng)",
    "Phường/Xã (Giao hàng)",
    "Số nhà, Đường phố (Giao hàng)",
    "Mã vùng (Giao hàng)",
    "Địa chỉ (Giao hàng)",
    "Tài khoản ngân hàng",
    "Mở tại ngân hàng",
    "Ngày thành lập/Ngày sinh",
    "Là khách hàng từ",
    "Doanh thu",
    "Quy mô nhân sự",
    "Loại hạn mức nợ",
    "Website",
    "Loại điều khoản TT",
    "Số nợ tối đa",
    "Số ngày/Ngày trong tháng",
    "Mô tả",
    "Chủ sở hữu",
    "Ngừng theo dõi",
    "Dùng chung",
    "Đối tác/CTV giới thiệu",
    "Là KH cá nhân",
    "Là đối tác/cộng tác viên",
    "Giới tính",
    "Ngày cấp",
    "Nơi cấp",
    "Mã số ĐVQHNS",
    "Fax",
    "Đơn vị chủ quản",
    "Số hộ chiếu",
    "Hình Thức thanh toán",
    "Xếp hạng khách hàng",
    "Là nhà phân phối",
]

CONTACT_HEADERS = [
    "Mã liên hệ (*)",
    "Xưng hô",
    "Họ và đệm",
    "Tên (*)",
    "Chức danh",
    "Phân cụm HĐ",
    "Tổ chức (*)",
    "Phòng ban",
    "ĐT cơ quan",
    "ĐT di động",
    "Điện thoại khác",
    "Email cá nhân",
    "Email cơ quan",
    "Nguồn gốc",
    "Nhân Viên Bán Hàng",
    "Quốc gia",
    "Tỉnh/Thành phố",
    "Quận/Huyện",
    "Phường/Xã",
    "Số nhà, Đường phố",
    "Mã vùng",
    "Địa chỉ",
    "Quốc gia (Giao hàng)",
    "Tỉnh/Thành phố (Giao hàng)",
    "Quận/Huyện (Giao hàng)",
    "Phường/Xã (Giao hàng)",
    "Số nhà, Đường phố (Giao hàng)",
    "Mã vùng (Giao hàng)",
    "Địa chỉ (Giao hàng)",
    "Ngày sinh",
    "Giới tính",
    "Tình trạng hôn nhân",
    "Facebook",
    "Tài khoản ngân hàng",
    "Mở tại ngân hàng",
    "Mô tả",
    "Dùng chung",
    "Ngừng theo dõi",
    "Phân loại khách hàng",
    "Là khách hàng từ",
    "Không gọi điện",
    "Không gửi Email",
    "Zalo",
]

ERROR_HEADERS = ["Row Number", "Error", "Original Data"]

# ─────────────────────────────────────────────────────────────────────────────
# Mapping functions
# ─────────────────────────────────────────────────────────────────────────────

def map_customer(first_row: Dict[str, Any], employee_names: List[str]) -> Dict[str, Any]:
    """
    Tạo 1 dòng customer MISA từ dòng đầu tiên của khách hàng trong Bravo.

    employee_names: danh sách tên nhân viên unique đã thu thập từ tất cả
                   các dòng giao hàng của khách hàng này.
    """
    return {
        "Mã khách hàng (*)":     _s(first_row.get("Mã khách hàng")),
        "Tên khách hàng (*)":    _s(first_row.get("Tên khách hàng")),
        "Email":                 _s(first_row.get("Email")),
        "Điện thoại":            _s(first_row.get("Điện thoại")),
        "Mã số thuế":            _s(first_row.get("MST")),
        # Địa chỉ GPKD (giấy phép kinh doanh) → Địa chỉ hóa đơn
        "Địa chỉ (Hóa đơn)":    _s(first_row.get("Địa chỉ GPKD")),
        # Địa chỉ liên hệ → Địa chỉ giao hàng cấp customer
        "Địa chỉ (Giao hàng)":  _s(first_row.get("Địa chỉ liên hệ")),
        "Loại khách hàng":       _s(first_row.get("Kênh bán hàng")),
        # Ghép tất cả tên nhân viên unique bằng dấu chấm phẩy
        "Nhân viên bán hàng":    "; ".join(employee_names) if employee_names else None,
        "Hình Thức thanh toán":  _s(first_row.get("Hình thức thanh toán")),
        "Loại điều khoản TT":    _s(first_row.get("Thời hạn thanh toán")),
    }


def map_contact(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Tạo 1 dòng contact MISA từ 1 dòng giao hàng Bravo.

    Mỗi dòng Bravo có 1 Mã giao hàng → 1 liên hệ MISA.
    Trả về None nếu dòng không có Mã giao hàng (bỏ qua).
    """
    ma_giao_hang = _s(row.get("Mã giao hàng"))
    if not ma_giao_hang:
        return None

    return {
        "Mã liên hệ (*)":                ma_giao_hang,
        # Người nhận là họ tên đầy đủ – để vào ô Tên (*)
        "Tên (*)":                        _s(row.get("Người nhận")),
        # Liên kết liên hệ về đúng khách hàng
        "Tổ chức (*)":                    _s(row.get("Mã khách hàng")),
        "ĐT di động":                     _s(row.get("Số điện thoại")),
        # Email của khách hàng dùng chung cho tất cả liên hệ
        "Email cơ quan":                  _s(row.get("Email")),
        "Quốc gia":                       "Việt Nam",
        # Địa chỉ GPKD → Địa chỉ (địa chỉ chính của liên hệ)
        "Địa chỉ":                        _s(row.get("Địa chỉ GPKD")),
        "Quốc gia (Giao hàng)":           "Việt Nam",
        "Tỉnh/Thành phố (Giao hàng)":     _s(row.get("Tỉnh/TP")),
        "Quận/Huyện (Giao hàng)":         _s(row.get("Quận/Huyện")),
        "Địa chỉ (Giao hàng)":            _s(row.get("Địa chỉ giao hàng")),
        "Nhân Viên Bán Hàng":             _s(row.get("Tên nhân viên")),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Result object
# ─────────────────────────────────────────────────────────────────────────────

class ConversionResult:
    def __init__(self) -> None:
        self.total_rows    = 0
        self.customer_rows = 0
        self.contact_rows  = 0
        self.error_rows    = 0
        self.elapsed_seconds = 0.0
        self.output_customer = ""
        self.output_contact  = ""
        self.output_error    = ""

# ─────────────────────────────────────────────────────────────────────────────
# Main convert function
# ─────────────────────────────────────────────────────────────────────────────

def convert(
    input_file: str,
    output_folder: str,
    log_fn: Callable[[str], None],
    progress_fn: Callable[[int, int], None],
    debug_mode: bool = False,
    pause_fn: Optional[Callable[[], None]] = None,
    stop_check: Optional[Callable[[], bool]] = None,
) -> ConversionResult:
    """
    Chuyển đổi file Bravo → customer.xlsx + contact.xlsx + error.xlsx.

    Tham số callback (chạy từ worker thread, không được gọi Tkinter trực tiếp):
      log_fn(msg)          – ghi log
      progress_fn(done, total) – cập nhật thanh tiến độ
      pause_fn()           – chỉ gọi khi debug_mode=True; blocking cho đến khi
                             user bấm Tiếp tục
      stop_check()         – trả về True nếu user bấm Dừng
    """
    result    = ConversionResult()
    t_start   = datetime.now()
    out_dir   = Path(output_folder)
    out_dir.mkdir(parents=True, exist_ok=True)

    customer_path = out_dir / "customer.xlsx"
    contact_path  = out_dir / "contact.xlsx"
    error_path    = out_dir / "error.xlsx"

    # ── Bước 1: Đọc headers + đếm dòng (pass đầu, nhanh) ─────────────────
    log_fn("Đang quét file...")
    bravo_headers, total_rows = _scan_file(input_file)
    log_fn(f"Tìm thấy {total_rows:,} dòng dữ liệu.")

    # ── Bước 2: Chuẩn bị workbook output ─────────────────────────────────
    # Contact → write-only (ghi ngay từng dòng để tiết kiệm RAM)
    contact_wb = Workbook(write_only=True)
    contact_ws = contact_wb.create_sheet("Nhập khẩu Liên hệ")
    contact_ws.append(CONTACT_HEADERS)

    # Error → write-only
    error_wb = Workbook(write_only=True)
    error_ws = error_wb.create_sheet("Lỗi")
    error_ws.append(ERROR_HEADERS)

    # Customer → gom vào dict trước (mỗi khách hàng có nhiều dòng giao hàng)
    # key = Mã khách hàng  →  (first_row, [employee_names])
    customers: Dict[str, Tuple[Dict[str, Any], List[str]]] = {}

    # ── Bước 3: Đọc từng dòng, mapping ───────────────────────────────────
    processed   = 0
    batch_start = datetime.now()

    wb_in = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws_in = wb_in.worksheets[0]

    try:
        for row_idx, raw_row in enumerate(ws_in.iter_rows(values_only=True), start=1):
            if row_idx < BRAVO_DATA_ROW:
                continue
            if not any(v is not None for v in raw_row):
                continue  # bỏ dòng trống

            if stop_check and stop_check():
                log_fn("⛔ Đã dừng theo yêu cầu.")
                break

            # Map raw tuple → dict theo tên cột
            row: Dict[str, Any] = {
                bravo_headers[i]: (raw_row[i] if i < len(raw_row) else None)
                for i in range(len(bravo_headers))
                if bravo_headers[i]
            }

            try:
                _process_row(row, row_idx, customers, contact_ws, error_ws, result)
            except Exception as exc:
                result.error_rows += 1
                log_fn(f"❌ Dòng {row_idx}: {exc}")
                error_ws.append([row_idx, str(exc), str(raw_row)])

            processed += 1
            result.total_rows = processed
            progress_fn(processed, total_rows)

            # ── Báo cáo sau mỗi batch ─────────────────────────────────────
            if processed % BATCH_SIZE == 0:
                elapsed = (datetime.now() - batch_start).total_seconds()
                batch_no = processed // BATCH_SIZE
                log_fn(
                    f"📦 Batch {batch_no} | "
                    f"Đã xử lý: {processed:,} | "
                    f"Lỗi: {result.error_rows} | "
                    f"Thời gian: {elapsed:.1f}s"
                )
                batch_start = datetime.now()

                if debug_mode and pause_fn:
                    pause_fn()  # blocking – GUI mở khoá khi user bấm Tiếp tục

    finally:
        wb_in.close()

    # ── Bước 4: Ghi customer.xlsx ─────────────────────────────────────────
    log_fn(f"Ghi {len(customers):,} khách hàng...")
    customer_wb = Workbook(write_only=True)
    customer_ws = customer_wb.create_sheet("Nhập khẩu Khách hàng")
    customer_ws.append(CUSTOMER_HEADERS)

    for first_row, emp_names in customers.values():
        mapped = map_customer(first_row, emp_names)
        customer_ws.append(_to_list(mapped, CUSTOMER_HEADERS))

    # ── Bước 5: Lưu tất cả ────────────────────────────────────────────────
    log_fn("Đang lưu file...")
    customer_wb.save(customer_path); customer_wb.close()
    contact_wb.save(contact_path);   contact_wb.close()
    error_wb.save(error_path);       error_wb.close()

    result.customer_rows     = len(customers)
    result.elapsed_seconds   = (datetime.now() - t_start).total_seconds()
    result.output_customer   = str(customer_path)
    result.output_contact    = str(contact_path)
    result.output_error      = str(error_path)
    return result

# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _process_row(
    row: Dict[str, Any],
    row_idx: int,
    customers: Dict[str, Tuple[Dict, List[str]]],
    contact_ws: Any,
    error_ws: Any,
    result: "ConversionResult",
) -> None:
    """Xử lý 1 dòng: cập nhật dict customers + ghi contact ngay."""

    ma_kh = _s(row.get("Mã khách hàng"))

    # ── Customer ──────────────────────────────────────────────────────────
    if ma_kh:
        if ma_kh not in customers:
            # Lần đầu gặp khách hàng này → lưu dòng đầu tiên
            customers[ma_kh] = (row, [])

        # Thu thập tên nhân viên unique cho khách hàng này
        nv = _s(row.get("Tên nhân viên"))
        if nv and nv not in customers[ma_kh][1]:
            customers[ma_kh][1].append(nv)

    # ── Contact ───────────────────────────────────────────────────────────
    contact = map_contact(row)
    if contact:
        contact_ws.append(_to_list(contact, CONTACT_HEADERS))
        result.contact_rows += 1


def _scan_file(path: str) -> Tuple[List[str], int]:
    """Pass 1: lấy headers + đếm tổng số dòng dữ liệu."""
    headers: List[str] = []
    count = 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == BRAVO_HEADER_ROW:
                headers = [str(v).strip() if v is not None else "" for v in row]
            if idx >= BRAVO_DATA_ROW and any(v is not None for v in row):
                count += 1
    finally:
        wb.close()
    return headers, count


def _to_list(data: Dict[str, Any], headers: List[str]) -> List[Any]:
    """Chuyển dict → list theo đúng thứ tự headers."""
    return [data.get(h) for h in headers]


def _s(v: Any) -> Optional[str]:
    """Trả về str đã strip, hoặc None nếu rỗng."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None
